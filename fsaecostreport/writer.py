#!/usr/bin/env python
"""
================================================================================
:mod:`writer` -- LaTeX writer of the cost report
================================================================================

.. module:: writer
   :synopsis: LaTeX writer of the cost report

.. inheritance-diagram:: fsaecostreport.writer

"""

# Script information for the file.
__author__ = "Philippe T. Pinard"
__email__ = "philippe.pinard@gmail.com"
__version__ = "0.1"
__copyright__ = "Copyright (c) 2011 Philippe T. Pinard"
__license__ = "GPL v3"

# Standard library modules.
import operator
import posixpath
import os.path
import csv

# Third party modules.
from openpyxl import Workbook
from openpyxl.style import NumberFormat, Fill, Color

# Local modules.
from fsaecostreport.latex import (
    create_tabular,
    escape as e,
    AuxReader,
    escape_math as m,
)
from fsaecostreport.component import Part, Assembly
import fsaecostreport.graph as graph

# Globals and constants variables.
from constants import DRAWINGS_DIR, PICTURES_DIR, LOGO_FILE

SUBTOTAL = operator.attrgetter("subtotal")


def decimal(number):
    if number < 0.01:
        return "%.2e" % number
    else:
        return "%4.2f" % number


def capitalize(value):
    if value:
        return value[0].upper() + value[1:]
    else:
        return value


def humanjoin(values, sort=True, andchr="and"):
    if not values:
        return ""
    elif len(values) == 1:
        return values[0]
    elif len(values) == 2:
        if sort:
            values.sort()
        return "%s %s %s" % (values[0], andchr, values[1])
    else:
        if sort:
            values.sort()
        return ", ".join(values[:-1]) + " " + andchr + " " + values[-1]


def _create_bom_row(component):
    if len(component.drawings) == 1:
        drawings = r"\pageref{dwg:%s-0}" % component.pn
    elif len(component.drawings) > 1:
        drawings = r"\pageref{dwg:%s-0}--\pageref{dwg:%s-%i}" % (
            component.pn,
            component.pn,
            len(component.drawings) - 1,
        )
    else:
        drawings = ""

    if len(component.pictures) == 1:
        pictures = r"\pageref{img:%s-0}" % component.pn
    elif len(component.pictures) > 1:
        pictures = r"\pageref{img:%s-0}--\pageref{img:%s-%i}" % (
            component.pn,
            component.pn,
            len(component.pictures) - 1,
        )
    else:
        pictures = ""

    # use tablecost instead of unitcost not to include the cost of parts
    unitcost = component.tablecost
    quantity = component.quantity
    totalcost = unitcost * quantity

    names = [
        e(capitalize(parent.name)) for parent in reversed(sorted(component.parents))
    ]
    assembly = humanjoin(names, andchr=r"\&")

    if isinstance(component, Assembly):
        row = [r"\hline\rowcolor[gray]{.9}{%s}" % e(capitalize(component.name))]
    elif isinstance(component, Part):
        row = [r"%s" % e(capitalize(component.name))]

    row += [
        r"\centering %s" % component.pn_base,
        r"\centering %s" % component.revision,
        r"\centering %s" % assembly,
        r"\raggedright %s" % e(capitalize(component.details)),
        r"\centering %i" % quantity,
        r"\raggedleft\$ %4.2f" % unitcost,
        r"\raggedleft\$ %4.2f" % totalcost,
        r"\centering\pageref{ct:%s}" % component.pn,
        r"\centering%s" % drawings,
        r"\centering%s" % pictures,
    ]

    return row


class CostReportLaTeXWriter(object):
    def write(self, basepath, metadata):
        lines = self._write(basepath, metadata)

        filename = metadata.filename + ".tex"
        with open(os.path.join(basepath, filename), "w") as out:
            for line in lines:
                out.write(line + "\n")

    def _write(self, basepath, metadata):
        lines = []

        lines += self.write_header(metadata)
        lines += [""]

        lines += [r"\begin{document}"]
        lines += [""]

        lines += self.write_fancy_header(metadata)
        lines += [""]
        lines += self.write_renewcommand()
        lines += [""]
        lines += self.write_colors(metadata)
        lines += [""]

        # content
        lines += self.write_frontmatter(basepath, metadata)
        lines += [""]

        lines += self.write_systems(metadata)
        lines += [""]

        lines += self.write_backmatter()
        lines += [""]

        lines += [r"\end{document}"]

        return lines

    def write_header(self, metadata):
        lines = []

        lines += [r"\documentclass[letterpaper,landscape]{report}"]

        lines += [r"\usepackage[scaled]{helvet}"]
        lines += [r"\renewcommand*\familydefault{\sfdefault}"]
        lines += [r"\usepackage[T1]{fontenc}"]
        lines += [r"\usepackage[top=3cm, bottom=3cm, right=1cm, left=1cm]{geometry}"]
        lines += [r"\usepackage{graphicx}"]
        lines += [r"\usepackage{multirow}"]
        lines += [r"\usepackage{url}"]
        lines += [r"\usepackage{amsmath}"]
        lines += [r"\usepackage{longtable}"]
        lines += [r"\usepackage{titlesec}"]
        lines += [r"\usepackage{array}"]
        lines += [r"\usepackage{colortbl}"]
        lines += [r"\usepackage{multicol}"]
        lines += [r"\usepackage{setspace}"]
        lines += [r"\usepackage[final]{pdfpages}"]
        lines += [r"\usepackage[english]{babel}"]
        lines += [r"\usepackage[latin1]{inputenc}"]
        lines += [r"\usepackage{fancyhdr}"]
        lines += [
            r"\usepackage[pdftitle={Cost Report %i}, " % metadata.year,
            r"pdfsubject={%s}, " % metadata.competition_name,
            r"pdfauthor={%s}, " % metadata.team_name,
            r"colorlinks=true, ",
            r"linkcolor=blue, ",
            r"pdfborder = 0 0 0, ",
            r"pdfhighlight = /I, ",
            r"pdfpagelabels]{hyperref}",
        ]

        return lines

    def write_fancy_header(self, metadata):
        team_name = metadata.team_name
        year = metadata.year

        lines = []

        lines += [r"\pagestyle{fancy}"]
        lines += [r"\fancyhf{}"]
        lines += [
            r"\lhead{\includegraphics[height=0.25cm]{%s}\hspace{10pt} %s -- %i Cost Report}"
            % (LOGO_FILE, team_name, year)
        ]
        lines += [r"\lfoot{\small\nouppercase{\leftmark}}"]
        lines += [r"\rfoot{\thepage}"]
        lines += [r"\fancypagestyle{plain}{"]
        lines += [r"\fancyhf{}"]
        lines += [
            r"\lhead{\includegraphics[height=0.25cm]{%s}\hspace{10pt} %s -- %i Cost Report}"
            % (LOGO_FILE, team_name, year)
        ]
        lines += [r"\lfoot{\small\nouppercase{\leftmark}}"]
        lines += [r"\rfoot{\thepage}}"]

        return lines

    def write_renewcommand(self):
        lines = []

        lines += [r"\renewcommand{\chaptername}{\sffamily System}"]
        lines += [r"\renewcommand{\thechapter}{\Alph{chapter}}"]
        lines += [r"\titleformat*{\section}{\Large\sffamily\raggedright}"]
        lines += [
            r"\renewcommand{\chaptermark}[1]{\markboth{\chaptername\ \thechapter:\ #1}{}}"
        ]
        lines += [r"\titlespacing{\subsubsection}{0pt}{*1}{*-1}"]

        return lines

    def write_colors(self, metadata):
        lines = []

        for system in metadata.systems:
            r = system.colour[0] / 255.0
            g = system.colour[1] / 255.0
            b = system.colour[2] / 255.0
            lines += [r"\definecolor{color%s}{rgb}{%f,%f,%f}" % (system.label, r, g, b)]

        return lines

    def write_frontmatter(self, basepath, metadata):
        lines = []

        lines += [r"\pagenumbering{roman}"]
        lines += [""]

        lines += self.write_toc()
        lines += [r"\newpage", ""]

        lines += self.write_introduction(metadata)
        lines += [r"\newpage", ""]

        lines += self.write_cost_summary(basepath, metadata)
        lines += [r"\newpage", ""]

        lines += self.write_standard_partnumbering(basepath)
        lines += [r"\newpage", ""]

        lines += self.write_sae_parts_bom(metadata)
        lines += [r"\newpage", ""]

        lines += [r"\pagenumbering{arabic}"]

        return lines

    def write_introduction(self, metadata):
        lines = []

        lines += [r"\section{Introduction}"]
        lines += [r"\doublespacing"]
        lines += [r"\begin{multicols}{2}"]

        lines += metadata.introduction

        lines += [r"\end{multicols}"]
        lines += [r"\singlespacing"]

        return lines

    def write_cost_summary(self, basepath, metadata):
        lines = []

        lines += [r"\section{Cost Summary}"]
        lines += [r"\renewcommand{\arraystretch}{1.5}"]

        data = self._create_cost_summary_lines(metadata)
        lines += create_tabular(
            data,
            environment="longtable",
            tableparameters="l",
            tablespec=r"p{20em} | p{8em} | p{8em} | p{8em} | p{8em} | p{12em}",
            format_before_tabular=r"\rowcolor[gray]{0}",
            format_after_header=r"\hline\endhead",
            format_between_rows=r"\hline",
            header_endrow=1,
        )
        lines += [r"\renewcommand{\arraystretch}{1}"]
        lines += [r"\newpage"]

        self._create_cost_summary_chart(basepath, metadata)
        lines += [r"\begin{center}"]
        lines += [r"\includegraphics[height=0.8\textheight]{cost_summary}"]
        lines += [r"\end{center}"]

        return lines

    def _create_cost_summary_lines(self, metadata):
        rows = []

        header = [
            r"\color{white} System",
            r"\color{white}\centering Materials",
            r"\color{white}\centering Processes",
            r"\color{white}\centering Fasteners",
            r"\color{white}\centering Tooling",
            r"\color{white}\centering Total",
        ]
        rows.append(header)

        materials_totalcost = 0.0
        processes_totalcost = 0.0
        fasteners_totalcost = 0.0
        toolings_totalcost = 0.0
        systems_totalcost = 0.0

        for system in metadata.systems:
            materials_cost = 0.0
            processes_cost = 0.0
            fasteners_cost = 0.0
            toolings_cost = 0.0
            system_cost = 0.0

            for component in system.get_components():
                qty = component.quantity

                materials_cost += sum(map(SUBTOTAL, component.materials)) * qty
                processes_cost += sum(map(SUBTOTAL, component.processes)) * qty
                fasteners_cost += sum(map(SUBTOTAL, component.fasteners)) * qty
                toolings_cost += sum(map(SUBTOTAL, component.toolings)) * qty

                # use of tablecost instead of unitcost not to include the cost
                # of parts twice in the system cost
                system_cost += component.tablecost * qty

            row = [
                r"\rowcolor{color%s}\raggedright %s" % (system.label, e(system.name)),
                r"\centering\$ %4.2f" % materials_cost,
                r"\centering\$ %4.2f" % processes_cost,
                r"\centering\$ %4.2f" % fasteners_cost,
                r"\centering\$ %4.2f" % toolings_cost,
                r"\centering\$ %4.2f" % system_cost,
            ]
            rows.append(row)

            materials_totalcost += materials_cost
            processes_totalcost += processes_cost
            fasteners_totalcost += fasteners_cost
            toolings_totalcost += toolings_cost
            systems_totalcost += system_cost

        row = [
            r"\hline\raggedright\textbf{ % s}" % "Total Vehicle",
            r"\centering\textbf{\$ %4.2f}" % materials_totalcost,
            r"\centering\textbf{\$ %4.2f}" % processes_totalcost,
            r"\centering\textbf{\$ %4.2f}" % fasteners_totalcost,
            r"\centering\textbf{\$ %4.2f}" % toolings_totalcost,
            r"\centering\textbf{\$ %4.2f}" % systems_totalcost,
        ]
        rows.append(row)

        return rows

    def _create_cost_summary_chart(self, basepath, metadata):
        graph.cost_summary(basepath, metadata)

    def write_standard_partnumbering(self, basepath):
        lines = []

        lines += [r"\section{Standard Part Numbering}"]

        path = os.path.join(basepath, "part_numbering.pdf")
        if os.path.exists(path):
            lines += [r"\begin{center}"]
            lines += [r"\includegraphics[height=0.8\textheight]{part_numbering}"]
            lines += [r"\end{center}"]

        return lines

    def write_sae_parts_bom(self, metadata):
        lines = []

        lines += [r"\section{SAE common parts}"]
        lines += [r"\noindent\emph{As per SAE Appendix C3}"]
        lines += [r"\renewcommand{\arraystretch}{1.1}"]

        data = self._write_sae_parts_bom_rows(metadata)
        lines += create_tabular(
            data,
            environment="longtable",
            tableparameters="l",
            tablespec=r"p{13em} | p{3em} | p{2em} | p{10em} | p{7em} | p{2em} | p{4.5em} | p{4.5em} | p{5em} | p{5em} | p{5em}",
            format_before_tabular=r"\rowcolor[gray]{0}",
            format_after_header=r"\hline\endhead",
            format_between_rows=r"\hline",
            header_endrow=1,
        )

        lines += [r"\renewcommand{\arraystretch}{1}"]

        return lines

    def _write_sae_parts_bom_rows(self, metadata):
        rows = []

        header = [
            r"\color{white} Component",
            r"\color{white}\centering Asm / Prt \#",
            r"\color{white}\centering Rev.",
            r"\color{white}\centering Assembly",
            r"\color{white} Description",
            r"\color{white}\centering Qty",
            r"\color{white}\centering Unit\\ Cost",
            r"\color{white}\centering Cost",
            r"\color{white}\centering Cost\\ Table",
            r"\color{white}\centering Drawing(s)",
            r"\color{white}\centering Photo(s)",
        ]
        rows.append(header)

        for system in metadata.systems:
            row = r"\multicolumn{11}{l}{\cellcolor{color%s}\textbf{%s}}" % (
                system.label,
                e(system.name),
            )
            rows.append([row])

            for component_name, pn in metadata.sae_parts.get(system, []):
                if pn:
                    component = system.get_component(pn)
                    row = _create_bom_row(component)
                else:
                    row = [
                        r"%s" % e(capitalize(component_name)),
                        r"\multicolumn{10}{l}{\emph{%s}}" % e("Not available"),
                    ]

                rows.append(row)

        return rows

    def write_toc(self):
        lines = []

        lines += [r"\setcounter{tocdepth}{1}"]
        lines += [r"\tableofcontents"]
        lines += [r"\pdfbookmark[0]{Contents}{contents}"]
        lines += [r"\newpage"]
        lines += [r"\setcounter{tocdepth}{4}"]
        lines += [""]

        return lines

    def write_systems(self, metadata):
        lines = []

        for system in metadata.systems:
            lines += SystemLaTeXWriter().write(system)
            lines += [""]

        return lines

    def write_backmatter(self):
        lines = []

        lines += [r"\renewcommand{\listfigurename}{List of Drawings}"]
        lines += [r"\pdfbookmark[0]{List of Drawings}{listofdd}"]
        lines += [r"\listoffigures"]

        return lines


class SystemLaTeXWriter(object):
    def write(self, system):
        hierarchy = system.get_hierarchy()

        lines = []

        lines += [r"\chapter{%s}" % e(system.name)]
        lines += [r"\newpage", ""]

        # BOM
        lines += self.write_bom(system, hierarchy)
        lines += [r"\newpage", ""]

        # cost tables
        lines += self.write_costtables(system, hierarchy)
        lines += [r"\newpage", ""]

        # drawings
        lines += self.write_drawings(system, hierarchy)
        lines += [r"\newpage", ""]

        # pictures
        lines += self.write_pictures(system, hierarchy)
        lines += [r"\newpage", ""]

        return lines

    def write_bom(self, system, hierarchy):
        lines = []

        lines += [r"\section{BOM}"]
        lines += [r"\renewcommand{\arraystretch}{1.1}"]

        data = self._create_bom_lines(system, hierarchy)
        lines += create_tabular(
            data,
            environment="longtable",
            tableparameters="l",
            tablespec=r"p{13em} | p{3em} | p{2em} | p{10em} | p{7em} | p{2em} | p{4.5em} | p{4.5em} | p{5em} | p{5em} | p{5em}",
            format_before_tabular=r"\rowcolor[gray]{0}",
            format_after_header=r"\hline\endhead",
            format_between_rows=r"\hline",
            header_endrow=1,
        )

        lines += [r"\renewcommand{\arraystretch}{1}"]

        return lines

    def _create_bom_lines(self, system, hierarchy):
        rows = []

        header = [
            r"\color{white} Component",
            r"\color{white}\centering Asm / Prt \#",
            r"\color{white}\centering Rev.",
            r"\color{white}\centering Assembly",
            r"\color{white} Description",
            r"\color{white}\centering Qty",
            r"\color{white}\centering Unit\\ Cost",
            r"\color{white}\centering Cost",
            r"\color{white}\centering Cost\\ Table",
            r"\color{white}\centering Drawing(s)",
            r"\color{white}\centering Photo(s)",
        ]
        rows.append(header)

        for component in hierarchy:
            row = _create_bom_row(component)
            rows.append(row)

        return rows

    def write_costtables(self, system, hierarchy):
        lines = []

        lines += [r"\section{Cost Tables}"]

        for component in hierarchy:
            if isinstance(component, Assembly):
                lines += AssemblyLaTeXWriter().write_costtables(component)
            elif isinstance(component, Part):
                lines += PartLaTeXWriter().write_costtables(component)

            lines += [r"\newpage"]

        return lines

    def write_drawings(self, system, hierarchy):
        lines = []

        lines += [r"\section{Technical Drawings}"]
        lines += ["The technical drawings are in the following pages."]

        for component in hierarchy:
            name = component.name.replace(",", "")
            pn = component.pn

            for index, drawing in enumerate(component.drawings):
                path = posixpath.join(
                    system.label, DRAWINGS_DIR, os.path.basename(drawing)
                )
                lines += [
                    r"\includepdf[pages={1}, addtolist={1,figure,%s (%s),dwg:%s-%i}]{%s}"
                    % (name, pn, pn, index, path)
                ]
                lines += [r"\addcontentsline{toc}{subsection}{%s (%s)}" % (name, pn)]

        return lines

    def write_pictures(self, system, hierarchy):
        lines = []

        lines += [r"\section{Pictures}"]

        for component in hierarchy:
            name = component.name
            pn = component.pn

            for index, picture in enumerate(component.pictures):
                path = posixpath.join(
                    system.label, PICTURES_DIR, os.path.basename(picture)
                )
                lines += [r"\subsection{%s (%s)}" % (name, pn)]
                lines += [r"\label{img:%s-%i}" % (pn, index)]
                lines += [r"\begin{center}"]
                lines += [r"\includegraphics[height=0.8\textheight]{%s}" % path]
                lines += [r"\end{center}"]
                lines += [r"\newpage"]

        return lines


class _ComponentLaTeXWriter(object):
    def write_costtables(self, component):
        lines = []

        lines += self._create_header_lines(component)

        lines += self.write_materials(component.materials)
        lines += self.write_processes(component.processes)
        lines += self.write_fasteners(component.fasteners)
        lines += self.write_toolings(component.toolings)

        return lines

    def _create_header_lines(self, component):
        lines = []

        lines += [r"\subsection{%s (%s)}" % (e(component.name), component.pn)]
        lines += [r"\label{ct:%s}" % component.pn]

        return lines

    def write_materials(self, materials):
        lines = []

        if materials:
            lines += [r"\subsubsection*{Materials}"]
            lines += [r"\renewcommand{\arraystretch}{1.25}"]

            data = self._create_materials_rows(materials)
            lines += create_tabular(
                data,
                environment="longtable",
                tableparameters="l",
                tablespec=r"m{14em}|m{12em}|m{6em}|m{6em}|m{4.5em}|m{4.5em}|m{4.5em}",
                format_before_tabular=r"\rowcolor[gray]{0}",
                format_after_header=r"\hline\endhead",
                format_between_rows=r"\hline",
                header_endrow=1,
            )

            lines += [r"\renewcommand{\arraystretch}{1}"]

        lines += [""]

        return lines

    def _create_materials_rows(self, materials):
        rows = []

        # header
        header = [
            r"\color{white} Material",
            r"\color{white} Use",
            r"\color{white}\centering Size 1",
            r"\color{white}\centering Size 2",
            r"\color{white}\centering Unit Cost",
            r"\color{white}\centering Qty",
            r"\color{white}\centering Sub Total",
        ]
        rows.append(header)

        # rows
        for material in materials:
            if material.size1:
                size1 = "%s $%s$" % (decimal(material.size1), m(material.unit1))
            else:
                size1 = r"\ "

            if material.size2:
                size2 = "%s $%s$" % (decimal(material.size2), m(material.unit2))
            else:
                size2 = r"\ "

            row = [
                r"\raggedright %s" % e(capitalize(material.name)),
                r"\raggedright %s" % e(capitalize(material.use)),
                r"\centering %s" % size1,
                r"\centering %s" % size2,
                r"\raggedleft\$ %4.2f" % material.unitcost,
                r"\centering %.3f" % material.quantity,
                r"\raggedleft\$ %4.2f" % material.subtotal,
            ]
            rows.append(row)

        # total
        totalcost = sum(map(SUBTOTAL, materials))
        row = [
            r"\multicolumn{6}{r}{\textbf{Total}}",
            r"\raggedleft\textbf{\$ %4.2f}" % totalcost,
        ]
        rows.append(row)

        return rows

    def write_processes(self, processes):
        lines = []

        if processes:
            lines += [r"\subsubsection*{Processes}"]
            lines += [r"\renewcommand{\arraystretch}{1.25}"]

            data = self._create_processes_rows(processes)
            lines += create_tabular(
                data,
                environment="longtable",
                tableparameters="l",
                tablespec=r"m{14em}|m{12em}|m{6em}|m{6em}|m{4.5em}|m{4.5em}",
                format_before_tabular=r"\rowcolor[gray]{0}",
                format_after_header=r"\hline\endhead",
                format_between_rows=r"\hline",
                header_endrow=1,
            )

            lines += [r"\renewcommand{\arraystretch}{1}"]

        lines += [""]

        return lines

    def _create_processes_rows(self, processes):
        rows = []

        # header
        header = [
            r"\color{white} Process",
            r"\color{white} Use",
            r"\color{white}\centering Cost",
            r"\color{white}\centering Qty",
            r"\color{white}\centering Multiplier",
            r"\color{white}\centering Sub Total",
        ]
        rows.append(header)

        # rows
        for process in processes:
            unitcost = r"\$ %4.2f / $%s$" % (process.unitcost, m(process.unit))

            if process.multiplier is None:
                multiplier = 1.0
            else:
                multiplier = process.multiplier

            row = [
                r"\raggedright %s" % e(capitalize(process.name)),
                r"\raggedright %s" % e(capitalize(process.use)),
                r"\raggedleft %s" % unitcost,
                r"\centering %4.2f" % process.quantity,
                r"\centering %4.2f" % multiplier,
                r"\raggedleft\$ %4.2f" % process.subtotal,
            ]
            rows.append(row)

        # total
        totalcost = sum(map(SUBTOTAL, processes))
        row = [
            r"\multicolumn{5}{r}{\textbf{Total}}",
            r"\raggedleft\textbf{\$ %4.2f}" % totalcost,
        ]
        rows.append(row)

        return rows

    def write_fasteners(self, fasteners):
        lines = []

        if fasteners:
            lines += [r"\subsubsection*{Fasteners}"]
            lines += [r"\renewcommand{\arraystretch}{1.25}"]

            data = self._create_fasteners_rows(fasteners)
            lines += create_tabular(
                data,
                environment="longtable",
                tableparameters="l",
                tablespec=r"m{14em}|m{12em}|m{6em}|m{6em}|m{4.5em}|m{4.5em}|m{4.5em}",
                format_before_tabular=r"\rowcolor[gray]{0}",
                format_after_header=r"\hline\endhead",
                format_between_rows=r"\hline",
                header_endrow=1,
            )

            lines += [r"\renewcommand{\arraystretch}{1}"]

        lines += [""]

        return lines

    def _create_fasteners_rows(self, fasteners):
        rows = []

        # header
        header = [
            r"\color{white} Fastener",
            r"\color{white} Use",
            r"\color{white}\centering Size 1",
            r"\color{white}\centering Size 2",
            r"\color{white}\centering Cost",
            r"\color{white}\centering Qty",
            r"\color{white}\centering Sub Total",
        ]
        rows.append(header)

        # rows
        for fastener in fasteners:
            if fastener.size1 is not None:
                size1 = "%4.2f $%s$" % (fastener.size1, m(fastener.unit1))
            else:
                size1 = r"\ "
            if fastener.size2 is not None:
                size2 = "%4.2f $%s$" % (fastener.size2, m(fastener.unit2))
            else:
                size2 = r"\ "

            row = [
                r"\raggedright %s" % e(capitalize(fastener.name)),
                r"\raggedright %s" % e(capitalize(fastener.use)),
                r"\centering %s" % size1,
                r"\centering %s" % size2,
                r"\raggedleft\$ %4.2f" % fastener.unitcost,
                r"\centering %s" % fastener.quantity,
                r"\raggedleft\$ %4.2f" % fastener.subtotal,
            ]
            rows.append(row)

        # total
        totalcost = sum(map(SUBTOTAL, fasteners))
        row = [
            r"\multicolumn{6}{r}{\textbf{Total}}",
            r"\raggedleft\textbf{\$ %4.2f}" % totalcost,
        ]
        rows.append(row)

        return rows

    def write_toolings(self, toolings):
        lines = []

        if toolings:
            lines += [r"\subsubsection*{Tooling}"]
            lines += [r"\renewcommand{\arraystretch}{1.25}"]

            data = self._create_toolings_rows(toolings)
            lines += create_tabular(
                data,
                environment="longtable",
                tableparameters="l",
                tablespec=r"m{14em}|m{12em}|m{8em}|m{6em}|m{6em}|m{4.5em}",
                format_before_tabular=r"\rowcolor[gray]{0}",
                format_after_header=r"\hline\endhead",
                format_between_rows=r"\hline",
                header_endrow=1,
            )

            lines += [r"\renewcommand{\arraystretch}{1}"]

        lines += [""]

        return lines

    def _create_toolings_rows(self, toolings):
        rows = []

        # header
        header = [
            r"\color{white} Tooling",
            r"\color{white} Use",
            r"\color{white}\centering Unit Cost",
            r"\color{white}\centering Qty",
            r"\color{white}\centering PVF",
            r"\color{white}\centering Sub Total",
        ]
        rows.append(header)

        # rows
        for tooling in toolings:
            unitcost = r"\$ %4.2f / $%s$" % (tooling.unitcost, m(tooling.unit))

            row = [
                r"\raggedright %s" % e(capitalize(tooling.name)),
                r"\raggedright %s" % e(capitalize(tooling.use)),
                r"\raggedleft %s" % unitcost,
                r"\centering %s" % tooling.quantity,
                r"\centering %s" % tooling.pvf,
                r"\raggedleft\$ %4.2f" % tooling.subtotal,
            ]
            rows.append(row)

        # total
        totalcost = sum(map(SUBTOTAL, toolings))
        row = [
            r"\multicolumn{5}{r}{\textbf{Total}}",
            r"\raggedleft\textbf{\$ %4.2f}" % totalcost,
        ]
        rows.append(row)

        return rows


class AssemblyLaTeXWriter(_ComponentLaTeXWriter):
    def write_costtables(self, component):
        lines = []

        lines += self._create_header_lines(component)

        lines += self.write_parts(component.components)
        lines += self.write_materials(component.materials)
        lines += self.write_processes(component.processes)
        lines += self.write_fasteners(component.fasteners)
        lines += self.write_toolings(component.toolings)

        return lines

    def write_parts(self, parts):
        lines = []

        if parts:
            lines += [r"\subsubsection*{Parts}"]
            lines += [r"\renewcommand{\arraystretch}{1.25}"]

            data = self._create_parts_rows(parts)
            lines += create_tabular(
                data,
                environment="longtable",
                tableparameters="l",
                tablespec=r"m{20em}|m{8em}|m{6em}|m{4.5em}|m{7em}",
                format_before_tabular=r"\rowcolor[gray]{0}",
                format_after_header=r"\hline\endhead",
                format_between_rows=r"\hline",
                header_endrow=1,
            )

            lines += [r"\renewcommand{\arraystretch}{1}"]

        lines += [""]

        return lines

    def _create_parts_rows(self, parts):
        rows = []

        # header
        header = [
            r"\color{white} Part",
            r"\color{white}\centering Part Number",
            r"\color{white}\centering Part Cost",
            r"\color{white}\centering Qty",
            r"\color{white}\centering Sub Total",
        ]
        rows.append(header)

        # rows
        totalcost = 0.0

        for part, quantity in reversed(sorted(parts.iteritems())):
            subtotal = part.unitcost * quantity

            row = [
                r"\raggedright %s" % e(capitalize(part.name)),
                r"\centering %s" % part.pn,
                r"\raggedleft\$ %4.2f" % part.unitcost,
                r"\centering %s" % quantity,
                r"\raggedleft\$ %4.2f" % subtotal,
            ]
            rows.append(row)

            totalcost += subtotal

        # total
        row = [
            r"\multicolumn{4}{r}{\textbf{Total}}",
            r"\raggedleft\textbf{\$ %4.2f}" % totalcost,
        ]
        rows.append(row)

        return rows


class PartLaTeXWriter(_ComponentLaTeXWriter):
    pass


class eBOMWriter(object):
    def write(self, basepath, metadata):
        pagerefs = AuxReader().read(basepath)

        filepath = os.path.join(basepath, metadata.filename + ".csv")
        writer = csv.writer(open(filepath, "w"))

        rows = self._create_rows(metadata, pagerefs)
        for row in rows:
            writer.writerow(row)

    def _create_rows(self, metadata, pagerefs):
        rows = []

        # spreadsheet header
        row = ["Competition Code", "FSAEM"] + [""] * 13
        rows.append(row)

        row = ["Year", str(metadata.year)[2:]] + [""] * 13
        rows.append(row)

        row = ["Car #", str(metadata.car_number).zfill(3)] + [""] * 13
        rows.append(row)

        # empty row
        row = [""] * 15
        rows.append(row)

        # table header
        header = [
            "Line Num.",
            "Area of Commodity",
            "Asm/Prt #",
            "Rev. Lvl.",
            "Asm",
            "Component",
            "Description",
            "Unit Cost",
            "Quantity",
            "Material Cost",
            "Process Cost",
            "Fastener Cost",
            "Tooling Cost",
            "Total Cost",
            "Details Page Number",
        ]
        rows.append(header)

        materials_totalcost = 0.0
        processes_totalcost = 0.0
        fasteners_totalcost = 0.0
        toolings_totalcost = 0.0
        systems_totalcost = 0.0

        for system in metadata.systems:
            (
                system_rows,
                materials_cost,
                processes_cost,
                fasteners_cost,
                toolings_cost,
                system_cost,
            ) = self._create_system_rows(system, pagerefs)

            rows.extend(system_rows)

            materials_totalcost += materials_cost
            processes_totalcost += processes_cost
            fasteners_totalcost += fasteners_cost
            toolings_totalcost += toolings_cost
            systems_totalcost += system_cost

        # add top row since the vehicle cost is now known
        row = (
            ["University", metadata.university]
            + [""] * 10
            + ["Total Vehicle Cost", "", systems_totalcost]
        )
        rows.insert(0, row)

        # vehicle total row
        row = [
            "",
            "Vehicle Total",
            "",
            "",
            "",
            "Total",
            "",
            "",
            "",
            materials_totalcost,
            processes_totalcost,
            fasteners_totalcost,
            toolings_totalcost,
            systems_totalcost,
            "",
        ]
        rows.append(row)

        return rows

    def _create_system_rows(self, system, pagerefs):
        rows = []

        materials_totalcost = 0.0
        processes_totalcost = 0.0
        fasteners_totalcost = 0.0
        toolings_totalcost = 0.0
        system_totalcost = 0.0

        for line_num, component in enumerate(system.get_hierarchy()):
            # use tablecost instead of unitcost not to include the cost of parts
            unitcost = component.tablecost
            quantity = component.quantity
            totalcost = unitcost * quantity

            materials_cost = sum(map(SUBTOTAL, component.materials))
            materials_totalcost += materials_cost * quantity

            processes_cost = sum(map(SUBTOTAL, component.processes))
            processes_totalcost += processes_cost * quantity

            fasteners_cost = sum(map(SUBTOTAL, component.fasteners))
            fasteners_totalcost += fasteners_cost * quantity

            toolings_cost = sum(map(SUBTOTAL, component.toolings))
            toolings_totalcost += toolings_cost * quantity

            system_totalcost += totalcost

            names = [e(capitalize(parent.name)) for parent in component.parents]
            assembly = humanjoin(names, andchr=r"\&")

            row = [
                line_num + 1,
                str(system),
                component.pn_base,
                component.revision,
                assembly,
                capitalize(component.name),
                capitalize(component.details),
                unitcost,
                quantity,
                materials_cost,
                processes_cost,
                fasteners_cost,
                toolings_cost,
                totalcost,
                pagerefs.get(component.pn, ""),
            ]
            rows.append(row)

        # area total row
        row = [
            "",
            str(system),
            "",
            "",
            "",
            "Area Total",
            "",
            "",
            "",
            materials_totalcost,
            processes_totalcost,
            fasteners_totalcost,
            toolings_totalcost,
            system_totalcost,
            "",
        ]
        rows.append(row)

        return (
            rows,
            materials_totalcost,
            processes_totalcost,
            fasteners_totalcost,
            toolings_totalcost,
            system_totalcost,
        )


class FSGBOMWriter(object):
    def write(self, basepath, metadata):
        wb = Workbook()

        # create cost tables
        for system in metadata.systems:
            self.write_system(wb, system, metadata)

        # remove first sheet
        wb.remove_sheet(wb.worksheets[0])

        filename = metadata.filename + ".xlsx"
        wb.save(os.path.join(basepath, filename))

    def write_system(self, wb, system, metadata):
        hierarchy = system.get_hierarchy()

        for component in hierarchy:
            sheet = wb.create_sheet(title=component.partnumber)

            if isinstance(component, Part):
                self.write_costtable_part(sheet, component, system, metadata)
            elif isinstance(component, Assembly):
                self.write_costtable_assembly(sheet, component, system, metadata)

            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 30
            sheet.column_dimensions["C"].width = 12
            sheet.column_dimensions["D"].width = 12
            sheet.column_dimensions["E"].width = 12
            sheet.column_dimensions["F"].width = 12
            sheet.column_dimensions["G"].width = 12
            sheet.column_dimensions["H"].width = 12

    def write_costtable_part(self, sheet, component, system, metadata):
        row = self.write_header_part(sheet, component, system, metadata)
        row = self.write_table_materials(sheet, component, row) + 1
        row = self.write_table_processes(sheet, component, row) + 1
        row = self.write_table_fasteners(sheet, component, row) + 1
        row = self.write_table_toolings(sheet, component, row) + 1

    def write_costtable_assembly(self, sheet, component, system, metadata):
        row = self.write_header_assembly(sheet, component, system, metadata)
        row = self.write_table_parts(sheet, component, row) + 1
        row = self.write_table_materials(sheet, component, row) + 1
        row = self.write_table_processes(sheet, component, row) + 1
        row = self.write_table_fasteners(sheet, component, row) + 1
        row = self.write_table_toolings(sheet, component, row) + 1

    def _set_header_cell(self, cell, value):
        cell.value = value
        cell.style.font.bold = True
        cell.style.fill.fill_type = Fill.FILL_SOLID
        cell.style.fill.start_color = Color("FFC0C0C0")
        cell.style.fill.end_color = Color("FFC0C0C0")

    def _set_money_cell(self, cell, value):
        cell.value = value
        cell.style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_USD_SIMPLE

    def write_header_part(self, sheet, component, system, metadata):
        self._set_header_cell(sheet.cell("A1"), "University")
        sheet.cell("B1").value = metadata.university

        self._set_header_cell(sheet.cell("D1"), "Car #")
        sheet.cell("E1").value = metadata.car_number

        self._set_header_cell(sheet.cell("A2"), "System")
        sheet.cell("B2").value = system.name

        self._set_header_cell(sheet.cell("A3"), "Assembly")
        sheet.cell("B3").value = ", ".join(
            assembly.name for assembly in component.parents
        )

        self._set_header_cell(sheet.cell("A4"), "Part")
        sheet.cell("B4").value = component.name

        self._set_header_cell(sheet.cell("A5"), "P/N Base")
        sheet.cell("B5").value = component.pn_base

        self._set_header_cell(sheet.cell("A6"), "Suffix")
        sheet.cell("B6").value = component.revision

        self._set_header_cell(sheet.cell("A7"), "Details")
        sheet.cell("B7").value = component.details

        self._set_header_cell(sheet.cell("G1"), "Unit cost")
        self._set_money_cell(sheet.cell("H1"), component.unitcost)

        self._set_header_cell(sheet.cell("G2"), "Quantity")
        sheet.cell("H2").value = component.quantity

        self._set_header_cell(sheet.cell("G4"), "Total cost")
        self._set_money_cell(sheet.cell("H4"), component.unitcost * component.quantity)

        return 9

    def write_header_assembly(self, sheet, component, system, metadata):
        self._set_header_cell(sheet.cell("A1"), "University")
        sheet.cell("B1").value = metadata.university

        self._set_header_cell(sheet.cell("D1"), "Car #")
        sheet.cell("E1").value = metadata.car_number

        self._set_header_cell(sheet.cell("A2"), "System")
        sheet.cell("B2").value = system.name

        self._set_header_cell(sheet.cell("A3"), "Assembly")
        sheet.cell("B3").value = component.name

        self._set_header_cell(sheet.cell("A4"), "P/N Base")
        sheet.cell("B4").value = component.pn_base

        self._set_header_cell(sheet.cell("A5"), "Suffix")
        sheet.cell("B5").value = component.revision

        self._set_header_cell(sheet.cell("A6"), "Details")
        sheet.cell("B6").value = component.details

        self._set_header_cell(sheet.cell("G1"), "Unit cost")
        self._set_money_cell(sheet.cell("H1"), component.unitcost)

        self._set_header_cell(sheet.cell("G2"), "Table cost")
        self._set_money_cell(sheet.cell("H2"), component.tablecost)

        self._set_header_cell(sheet.cell("G3"), "Quantity")
        sheet.cell("H3").value = component.quantity

        self._set_header_cell(sheet.cell("G4"), "Total cost")
        self._set_money_cell(sheet.cell("H4"), component.unitcost * component.quantity)

        return 8

    def write_table_parts(self, sheet, component, row):
        sheet.cell(row=row, column=0).value = "Parts"
        sheet.cell(row=row, column=0).style.font.bold = True

        # header
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=0), "P/N")
        self._set_header_cell(sheet.cell(row=row, column=1), "Part")
        self._set_header_cell(sheet.cell(row=row, column=2), "Part cost")
        self._set_header_cell(sheet.cell(row=row, column=3), "Quantity")
        self._set_header_cell(sheet.cell(row=row, column=4), "Sub total")

        # values
        totalcost = 0.0
        parts = component.components
        for part, quantity in reversed(sorted(parts.iteritems())):
            row += 1

            subtotal = part.unitcost * quantity
            totalcost += subtotal

            sheet.cell(row=row, column=0).value = part.pn
            sheet.cell(row=row, column=1).value = part.name
            self._set_money_cell(sheet.cell(row=row, column=2), part.unitcost)
            sheet.cell(row=row, column=3).value = quantity
            self._set_money_cell(sheet.cell(row=row, column=4), subtotal)

        # total
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=3), "Sub total")
        self._set_header_cell(sheet.cell(row=row, column=4), totalcost)
        sheet.cell(
            row=row, column=4
        ).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_USD_SIMPLE

        return row + 1

    def write_table_materials(self, sheet, component, row):
        sheet.cell(row=row, column=0).value = "Materials"
        sheet.cell(row=row, column=0).style.font.bold = True

        # header
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=0), "Material")
        self._set_header_cell(sheet.cell(row=row, column=1), "Use")
        self._set_header_cell(sheet.cell(row=row, column=2), "Size 1")
        self._set_header_cell(sheet.cell(row=row, column=3), "Size 2")
        self._set_header_cell(sheet.cell(row=row, column=4), "Unit cost")
        self._set_header_cell(sheet.cell(row=row, column=5), "Quantity")
        self._set_header_cell(sheet.cell(row=row, column=6), "Sub total")

        # values
        materials = component.materials
        for material in materials:
            row += 1

            sheet.cell(row=row, column=0).value = material.name
            sheet.cell(row=row, column=1).value = material.use
            if material.size1:
                sheet.cell(row=row, column=2).value = "%s %s" % (
                    decimal(material.size1),
                    material.unit1,
                )
            if material.size2:
                sheet.cell(row=row, column=3).value = "%s %s" % (
                    decimal(material.size2),
                    material.unit2,
                )
            self._set_money_cell(sheet.cell(row=row, column=4), material.unitcost)
            sheet.cell(row=row, column=5).value = material.quantity
            self._set_money_cell(sheet.cell(row=row, column=6), material.subtotal)

        # total
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=5), "Sub total")
        self._set_header_cell(
            sheet.cell(row=row, column=6), sum(map(SUBTOTAL, materials))
        )
        sheet.cell(
            row=row, column=6
        ).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_USD_SIMPLE

        return row + 1

    def write_table_processes(self, sheet, component, row):
        sheet.cell(row=row, column=0).value = "Processes"
        sheet.cell(row=row, column=0).style.font.bold = True

        # header
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=0), "Process")
        self._set_header_cell(sheet.cell(row=row, column=1), "Use")
        self._set_header_cell(sheet.cell(row=row, column=2), "Cost")
        self._set_header_cell(sheet.cell(row=row, column=3), "Quantity")
        self._set_header_cell(sheet.cell(row=row, column=4), "Multiplier")
        self._set_header_cell(sheet.cell(row=row, column=5), "Sub total")

        # values
        processes = component.processes
        for process in processes:
            row += 1

            sheet.cell(row=row, column=0).value = process.name
            sheet.cell(row=row, column=1).value = process.use
            sheet.cell(row=row, column=2).value = "%4.2f / %s" % (
                process.unitcost,
                process.unit,
            )
            sheet.cell(row=row, column=3).value = process.quantity
            sheet.cell(row=row, column=4).value = process.multiplier or 1.0
            self._set_money_cell(sheet.cell(row=row, column=5), process.subtotal)

        # total
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=4), "Sub total")
        self._set_header_cell(
            sheet.cell(row=row, column=5), sum(map(SUBTOTAL, processes))
        )
        sheet.cell(
            row=row, column=5
        ).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_USD_SIMPLE

        return row + 1

    def write_table_fasteners(self, sheet, component, row):
        sheet.cell(row=row, column=0).value = "Fasteners"
        sheet.cell(row=row, column=0).style.font.bold = True

        # header
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=0), "Fastener")
        self._set_header_cell(sheet.cell(row=row, column=1), "Use")
        self._set_header_cell(sheet.cell(row=row, column=2), "Size 1")
        self._set_header_cell(sheet.cell(row=row, column=3), "Size 2")
        self._set_header_cell(sheet.cell(row=row, column=4), "Unit cost")
        self._set_header_cell(sheet.cell(row=row, column=5), "Quantity")
        self._set_header_cell(sheet.cell(row=row, column=6), "Sub total")

        # values
        fasteners = component.fasteners
        for fastener in fasteners:
            row += 1

            sheet.cell(row=row, column=0).value = fastener.name
            sheet.cell(row=row, column=1).value = fastener.use
            if fastener.size1:
                sheet.cell(row=row, column=2).value = "%s %s" % (
                    decimal(fastener.size1),
                    fastener.unit1,
                )
            if fastener.size2:
                sheet.cell(row=row, column=3).value = "%s %s" % (
                    decimal(fastener.size2),
                    fastener.unit2,
                )
            self._set_money_cell(sheet.cell(row=row, column=4), fastener.unitcost)
            sheet.cell(row=row, column=5).value = fastener.quantity
            self._set_money_cell(sheet.cell(row=row, column=6), fastener.subtotal)

        # total
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=5), "Sub total")
        self._set_header_cell(
            sheet.cell(row=row, column=6), sum(map(SUBTOTAL, fasteners))
        )
        sheet.cell(
            row=row, column=6
        ).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_USD_SIMPLE

        return row + 1

    def write_table_toolings(self, sheet, component, row):
        sheet.cell(row=row, column=0).value = "Tooling"
        sheet.cell(row=row, column=0).style.font.bold = True

        # header
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=0), "Tooling")
        self._set_header_cell(sheet.cell(row=row, column=1), "Use")
        self._set_header_cell(sheet.cell(row=row, column=2), "Unit cost")
        self._set_header_cell(sheet.cell(row=row, column=3), "Quantity")
        self._set_header_cell(sheet.cell(row=row, column=4), "PVF")
        self._set_header_cell(sheet.cell(row=row, column=5), "Sub total")

        # values
        toolings = component.toolings
        for tooling in toolings:
            row += 1

            sheet.cell(row=row, column=0).value = tooling.name
            sheet.cell(row=row, column=1).value = tooling.use
            sheet.cell(row=row, column=2).value = "%4.2f / %s" % (
                tooling.unitcost,
                tooling.unit,
            )
            sheet.cell(row=row, column=3).value = tooling.quantity
            sheet.cell(row=row, column=4).value = tooling.pvf
            self._set_money_cell(sheet.cell(row=row, column=5), tooling.subtotal)

        # total
        row += 1
        self._set_header_cell(sheet.cell(row=row, column=4), "Sub total")
        self._set_header_cell(
            sheet.cell(row=row, column=5), sum(map(SUBTOTAL, toolings))
        )
        sheet.cell(
            row=row, column=5
        ).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_USD_SIMPLE

        return row + 1


class FSGAppendixLaTeXWriter(CostReportLaTeXWriter):
    def write_header(self, metadata):
        lines = []

        lines += [r"\documentclass[letterpaper,landscape]{report}"]

        lines += [r"\usepackage[scaled]{helvet}"]
        lines += [r"\renewcommand*\familydefault{\sfdefault}"]
        lines += [r"\usepackage[T1]{fontenc}"]
        lines += [r"\usepackage[top=3cm, bottom=3cm, right=1cm, left=1cm]{geometry}"]
        lines += [r"\usepackage{graphicx}"]
        lines += [r"\usepackage{multirow}"]
        lines += [r"\usepackage{url}"]
        lines += [r"\usepackage{amsmath}"]
        lines += [r"\usepackage{longtable}"]
        lines += [r"\usepackage{titlesec}"]
        lines += [r"\usepackage{array}"]
        lines += [r"\usepackage{colortbl}"]
        lines += [r"\usepackage{multicol}"]
        lines += [r"\usepackage{setspace}"]
        lines += [r"\usepackage[final]{pdfpages}"]
        lines += [r"\usepackage[english]{babel}"]
        lines += [r"\usepackage[latin1]{inputenc}"]
        lines += [r"\usepackage{fancyhdr}"]
        lines += [
            r"\usepackage[pdftitle={Cost Report %i - Supporting material}, "
            % metadata.year,
            r"pdfsubject={%s}, " % metadata.competition_name,
            r"pdfauthor={%s}, " % metadata.team_name,
            r"colorlinks=true, ",
            r"linkcolor=blue, ",
            r"pdfborder = 0 0 0, ",
            r"pdfhighlight = /I, ",
            r"pdfpagelabels]{hyperref}",
        ]

        return lines

    def write_fancy_header(self, metadata):
        team_name = metadata.team_name
        year = metadata.year

        lines = []

        lines += [r"\pagestyle{fancy}"]
        lines += [r"\fancyhf{}"]
        lines += [
            r"\lhead{\includegraphics[height=0.25cm]{%s}\hspace{10pt} %s -- %i Cost Report -- Supporting Material}"
            % (LOGO_FILE, team_name, year)
        ]
        lines += [r"\lfoot{\small\nouppercase{\leftmark}}"]
        lines += [r"\rfoot{\thepage}"]
        lines += [r"\fancypagestyle{plain}{"]
        lines += [r"\fancyhf{}"]
        lines += [
            r"\lhead{\includegraphics[height=0.25cm]{%s}\hspace{10pt} %s -- %i Cost Report -- Supporting Material}"
            % (LOGO_FILE, team_name, year)
        ]
        lines += [r"\lfoot{\small\nouppercase{\leftmark}}"]
        lines += [r"\rfoot{\thepage}}"]

        return lines

    def write_sae_parts_bom(self, metadata):
        return []

    #        lines = []
    #
    #        lines += [r'\section{SAE common parts}']
    #        lines += [r'\noindent\emph{As per SAE Appendix C3}']
    #        lines += [r'\renewcommand{\arraystretch}{1.1}']
    #
    #        data = self._write_sae_parts_bom_rows(metadata)
    #        lines += \
    #            create_tabular(data, environment='longtable',
    #                               tableparameters='l',
    #                               tablespec=r'p{13em} | p{3em} | p{2em} | p{10em} | p{12em} | p{2em} | p{4.5em} | p{4.5em} | p{5em} | p{5em}',
    #                               format_before_tabular=r'\rowcolor[gray]{0}',
    #                               format_after_header=r'\hline\endhead',
    #                               format_between_rows=r'\hline', header_endrow=1)
    #
    #        lines += [r'\renewcommand{\arraystretch}{1}']
    #
    #        return lines
    #
    #    def _write_sae_parts_bom_rows(self, metadata):
    #        rows = []
    #
    #        header = [r'\color{white} Component',
    #                  r'\color{white}\centering Asm / Prt \#',
    #                  r'\color{white}\centering Rev.',
    #                  r'\color{white}\centering Assembly',
    #                  r'\color{white} Description',
    #                  r'\color{white}\centering Qty',
    #                  r'\color{white}\centering Unit\\ Cost',
    #                  r'\color{white}\centering Cost',
    #                  r'\color{white}\centering Drawing(s)',
    #                  r'\color{white}\centering Photo(s)']
    #        rows.append(header)
    #
    #        for system in metadata.systems:
    #            row = r'\multicolumn{10}{l}{\cellcolor{color%s}\textbf{%s}}' % \
    #                        (system.label, e(system.name))
    #            rows.append([row])
    #
    #            for component_name, pn in metadata.sae_parts.get(system, []):
    #                if pn:
    #                    component = system.get_component(pn)
    #                    row = _create_bom_row(component)
    #                    row.pop(8)
    #                else:
    #                    row = [r'%s' % e(capitalize(component_name)),
    #                           r'\multicolumn{9}{l}{\emph{%s}}' % e('Not available')]
    #
    #                rows.append(row)
    #
    #        return rows

    def write_systems(self, metadata):
        lines = []

        for system in metadata.systems:
            lines += self.write_system(system)
            lines += [""]

        return lines

    def write_system(self, system):
        lines = []

        lines += [r"\chapter{%s}" % e(system.name)]
        lines += [r"\newpage", ""]

        # Drawings and pictures
        hierarchy = system.get_hierarchy()
        for component in hierarchy:
            component_lines = self.write_component(system, component)
            if component_lines:
                lines += component_lines
                lines += [r"\newpage", ""]

        return lines

    def write_component(self, system, component):
        if not component.pictures and not component.drawings:
            return []

        name = component.name
        pn = component.pn

        lines = []

        if component.pictures:
            lines += [r"\section{%s (%s)}" % (e(name), pn)]

            for index, picture in enumerate(component.pictures):
                path = posixpath.join(
                    system.label, PICTURES_DIR, os.path.basename(picture)
                )
                lines += [r"\begin{center}"]
                lines += [r"\includegraphics[height=0.8\textheight]{%s}" % path]
                lines += [r"\label{img:%s-%i}" % (pn, index)]
                lines += [r"\end{center}"]
                lines += [r"\newpage"]

            for index, drawing in enumerate(component.drawings):
                path = posixpath.join(
                    system.label, DRAWINGS_DIR, os.path.basename(drawing)
                )
                lines += [
                    r"\includepdf[pages={1}, addtolist={1,figure,%s (%s),dwg:%s-%i}]{%s}"
                    % (e(name), pn, pn, index, path)
                ]
        #                lines += [r'\addcontentsline{toc}{subsection}{Drawing %i}' % (index + 1,)]
        else:
            for index, drawing in enumerate(component.drawings):
                path = posixpath.join(
                    system.label, DRAWINGS_DIR, os.path.basename(drawing)
                )
                lines += [
                    r"\includepdf[pages={1}, addtolist={1,figure,%s (%s),dwg:%s-%i}]{%s}"
                    % (e(name), pn, pn, index, path)
                ]
                #                lines += [r'\addcontentsline{toc}{subsection}{Drawing %i}' % (index + 1,)]
                if index == 0:
                    lines += [
                        r"\addcontentsline{toc}{section}{%s (%s)}" % (e(name), pn)
                    ]
        return lines

    def write_backmatter(self):
        return []
